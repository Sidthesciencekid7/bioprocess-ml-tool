import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import shap
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_absolute_error
from sklearn.model_selection import LeaveOneGroupOut, cross_val_predict
from xgboost import XGBRegressor
from fpdf import FPDF
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import io
import tempfile
import os
import google.generativeai as genai
from datetime import datetime

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(page_title="Bioprocess ML Reporter", page_icon="🧬", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    .main-header { background: linear-gradient(135deg, #1a3a4a 0%, #0d6e8a 100%); padding: 1.5rem 2rem; border-radius: 12px; margin-bottom: 1.5rem; color: white; }
    .main-header h1 { margin: 0; font-size: 2rem; }
    .main-header p  { margin: 0.3rem 0 0; opacity: 0.8; font-size: 0.95rem; }
    .step-card { background: #f8fafc; border: 1px solid #e2e8f0; border-left: 4px solid #0d6e8a; padding: 1rem 1.2rem; border-radius: 8px; margin-bottom: 1rem; }
    .flag-warn { background: #fff7ed; border-left: 4px solid #f97316; padding: 0.5rem 1rem; border-radius: 6px; margin: 0.3rem 0; }
    .flag-ok   { background: #f0fdf4; border-left: 4px solid #22c55e; padding: 0.5rem 1rem; border-radius: 6px; margin: 0.3rem 0; }
    .insight-box { background: #f0f9ff; border: 1px solid #bae6fd; border-left: 4px solid #0284c7; padding: 1rem 1.2rem; border-radius: 8px; margin: 0.8rem 0; line-height: 1.6; }
    .chart-insight { background: #fafafa; border: 1px solid #e5e7eb; padding: 0.7rem 1rem; border-radius: 6px; margin-top: 0.5rem; font-size: 0.9rem; color: #374151; }
    .warning-box { background: #fef3c7; border: 1px solid #fcd34d; padding: 0.8rem 1rem; border-radius: 8px; margin: 0.5rem 0; font-size: 0.9rem; }
    .ai-badge { display: inline-block; background: #4f46e5; color: white; font-size: 0.7rem; padding: 0.1rem 0.4rem; border-radius: 4px; margin-left: 0.4rem; vertical-align: middle; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# SESSION STATE
# =============================================================================
for key in ['df','feature_cols','target_col','time_col','group_col','rf','xgb',
            'y','rf_pred','xgb_pred','flags','df_plot','shap_values','X_model',
            'ai_summary','ai_flags','ai_titer','ai_feat','ai_corr','ai_shap',
            'ai_harvest','ai_feed','ai_cellline','ai_next_run']:
    if key not in st.session_state:
        st.session_state[key] = None

# =============================================================================
# SIDEBAR
# =============================================================================
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.markdown("---")
    st.markdown("### 🚩 Process Flag Thresholds")
    nh3_thresh = st.slider("NH3 max (mM)",            1.0, 20.0, 5.0,  0.5)
    glc_thresh = st.slider("Glucose min (mM)",         0.1,  5.0, 2.0,  0.1)
    vcd_thresh = st.slider("Min peak VCD (x10^6/mL)", 1.0, 50.0, 10.0, 1.0)
    lg_thresh  = st.slider("Lactate/Glucose ratio",    0.5,  5.0, 2.0,  0.1)
    st.markdown("---")
    st.markdown("### 🤖 Model Settings")
    n_estimators = st.slider("Number of trees", 50, 500, 150, 50)
    use_shap     = st.checkbox("Compute SHAP values", value=True)
    st.markdown("---")
    st.caption("Bioprocess ML Reporter v3.0")

st.markdown('<div class="main-header"><h1>🧬 Bioprocess ML Report Generator</h1><p>Upload your bioprocess data and get instant ML predictions, process flags, and AI-powered plain-English interpretation.</p></div>', unsafe_allow_html=True)

# =============================================================================
# AI HELPER — Gemini, fully internal, no user input needed
# =============================================================================
def call_ai(prompt):
    try:
        api_key = st.secrets.get("GEMINI_API_KEY", None)
        if not api_key:
            st.session_state['ai_error'] = "GEMINI_API_KEY not found in Streamlit secrets."
            return None
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.5-flash")
        response = model.generate_content(prompt)
        st.session_state['ai_error'] = None
        return response.text.strip()
    except Exception as e:
        st.session_state['ai_error'] = str(e)
        return None


def ai_interpret(prompt, fallback=None):
    result = call_ai(prompt)
    return result if result else fallback


def run_all_ai_interpretations(df_plot, target_col, time_col, feature_cols,
                                rf, y, rf_pred, xgb_pred, flags, group_col):
    warn_flags = [m for s, m in flags if s == "warn"]
    ok_flags   = [m for s, m in flags if s == "ok"]
    top_feats  = sorted(zip(feature_cols, rf.feature_importances_), key=lambda x: x[1], reverse=True)[:3]
    final_val  = float(df_plot[target_col].iloc[-1])
    peak_val   = float(df_plot[target_col].max())
    duration   = int(df_plot[time_col].max()) if time_col in df_plot.columns else "unknown"
    rf_r2      = r2_score(y, rf_pred)
    xgb_r2     = r2_score(y, xgb_pred)

    prompt = f"""You are an expert bioprocess scientist interpreting ML results for a mixed audience including non-specialists.

Data:
- Target: {target_col}, Final value: {final_val:.2f}, Peak: {peak_val:.2f}, Duration: {duration} units
- RF R2: {rf_r2:.3f}, XGBoost R2: {xgb_r2:.3f}
- Top 3 predictive features: {', '.join([f[0] for f in top_feats])}
- Warnings: {'; '.join(warn_flags) if warn_flags else 'None'}
- Positive indicators: {'; '.join(ok_flags) if ok_flags else 'None'}

Return ONLY a valid JSON object with exactly these keys (no markdown, no backticks, no extra text):
{{
  "summary": "3-4 sentence plain English run summary describing what happened and whether the outcome was good or bad",
  "flags": "One line per flag explaining what it means in plain English and why it matters",
  "titer": "2 sentences explaining what the titer curve shape tells us and what the gap between actual and predicted means",
  "features": "2-3 sentences explaining why the top 3 features matter biologically and what to watch in future runs",
  "correlation": "2 sentences explaining what the correlation heatmap shows and what action to take",
  "shap": "2 sentences explaining what SHAP adds beyond feature importance and why it is useful",
  "next_run": "3 specific numbered actionable recommendations for the next run mentioning specific features or thresholds"
}}"""

    response = call_ai(prompt)
    if not response:
        return

    try:
        import json
        # Strip any accidental markdown fences
        clean = response.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        data = json.loads(clean)
        st.session_state.ai_summary  = data.get("summary")
        st.session_state.ai_flags    = data.get("flags")
        st.session_state.ai_titer    = data.get("titer")
        st.session_state.ai_feat     = data.get("features")
        st.session_state.ai_corr     = data.get("correlation")
        st.session_state.ai_shap     = data.get("shap")
        st.session_state.ai_next_run = data.get("next_run")
    except Exception as e:
        # If JSON parsing fails, put the raw response in the summary
        st.session_state.ai_summary = response

# =============================================================================
# HELPERS
# =============================================================================
def clean_text(t):
    return (t.replace("⚠️","[!]").replace("✅","[OK]").replace("⚠","[!]")
             .replace("°","").replace("×","x").replace("⁶","6").replace("²","2")
             .replace("—","-").replace("–","-")
             .replace("\u2019","'").replace("\u2018","'")
             .replace("\u201c",'"').replace("\u201d",'"'))


def smart_detect(cols):
    target_hints = ['titer','igg','antibody','mab','product','titre','ethanol','biomass']
    time_hints   = ['day','time','hour','date','sample']
    group_hints  = ['id','experiment','run','batch','cell line','cellline']
    target = next((c for c in cols for h in target_hints if h.lower() in c.lower()), cols[0])
    time   = next((c for c in cols for h in time_hints   if h.lower() in c.lower()), cols[0])
    group  = next((c for c in cols for h in group_hints  if h.lower() in c.lower()), None)
    return target, time, group


@st.cache_data(show_spinner=False)
def load_excel(file_bytes, sheet_name):
    buf = io.BytesIO(file_bytes)
    df = pd.read_excel(buf, sheet_name=sheet_name)
    if df.iloc[0].apply(lambda x: isinstance(x, str)).sum() > len(df.columns) * 0.5:
        df.columns = df.iloc[0]
        df = df.drop(0).reset_index(drop=True)
    df = df.loc[:, ~df.columns.duplicated()]
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='ignore')
    return df


@st.cache_data(show_spinner=False)
def load_csv(file_bytes):
    buf = io.BytesIO(file_bytes)
    df = pd.read_csv(buf)
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='ignore')
    return df


def generate_flags(df, target_col, time_col, nh3_thresh, glc_thresh, vcd_thresh, lg_thresh):
    flags = []
    if 'NH3 (mM)' in df.columns:
        max_nh3 = df['NH3 (mM)'].max()
        if max_nh3 > nh3_thresh:
            day = df.loc[df['NH3 (mM)'].idxmax(), time_col] if time_col in df.columns else "?"
            flags.append(("warn", f"NH3 exceeded {nh3_thresh} mM (peak: {max_nh3:.2f} mM on {time_col}={day})"))
    if 'Glucose (mM)' in df.columns:
        min_glc = df['Glucose (mM)'].min()
        if min_glc < glc_thresh:
            day = df.loc[df['Glucose (mM)'].idxmin(), time_col] if time_col in df.columns else "?"
            flags.append(("warn", f"Glucose dropped below {glc_thresh} mM (min: {min_glc:.2f} mM on {time_col}={day})"))
    vcd_col = 'Viable Cell Concentration (10^6 cells/mL)'
    if vcd_col in df.columns:
        peak_vcd = df[vcd_col].max()
        if peak_vcd < vcd_thresh:
            flags.append(("warn", f"Peak VCD below {vcd_thresh}x10^6 cells/mL (peak: {peak_vcd:.2f})"))
    if 'Lactate (mM)' in df.columns and 'Glucose (mM)' in df.columns:
        ratio = df['Lactate (mM)'] / (df['Glucose (mM)'] + 1e-9)
        max_ratio = ratio.max()
        if max_ratio > lg_thresh:
            flags.append(("warn", f"Lactate/Glucose ratio exceeded {lg_thresh:.1f} (max: {max_ratio:.2f}) - possible metabolic shift"))
    numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
    for col in numeric_cols:
        if col in [target_col, time_col]:
            continue
        diff = df[col].diff().abs()
        mean_d, std_d = diff.mean(), diff.std()
        if std_d > 0 and (diff > mean_d + 3 * std_d).any():
            day = df.loc[diff.idxmax(), time_col] if time_col in df.columns else "?"
            flags.append(("warn", f"Spike detected in '{col}' near {time_col}={day}"))
    if df[target_col].is_monotonic_increasing:
        flags.append(("ok", f"Target '{target_col}' increases monotonically - healthy profile"))
    final = df[target_col].iloc[-1]
    peak  = df[target_col].max()
    if peak > 0 and final < 0.5 * peak:
        flags.append(("warn", f"Final {target_col} ({final:.2f}) is less than 50% of peak ({peak:.2f}) - possible crash or dilution"))
    missing = df.isnull().sum()
    bad_cols = missing[missing > 0]
    if len(bad_cols) > 0:
        flags.append(("warn", f"Missing values in: {', '.join(bad_cols.index.tolist()[:5])}"))
    if not flags:
        flags.append(("ok", "No anomalies detected - run profile looks healthy!"))
    return flags


@st.cache_data(show_spinner=False)
def train_models(X_vals, X_cols, y_vals, y_name, groups_list, n_estimators):
    X = pd.DataFrame(X_vals, columns=X_cols)
    y = pd.Series(y_vals, name=y_name)
    groups = pd.Series(groups_list) if groups_list is not None else None
    rf  = RandomForestRegressor(n_estimators=n_estimators, random_state=42, n_jobs=-1)
    xgb = XGBRegressor(n_estimators=n_estimators, learning_rate=0.05, random_state=42, verbosity=0, n_jobs=-1)
    if groups is not None and groups.nunique() >= 3:
        logo     = LeaveOneGroupOut()
        rf_pred  = cross_val_predict(rf,  X, y, cv=logo, groups=groups)
        xgb_pred = cross_val_predict(xgb, X, y, cv=logo, groups=groups)
    else:
        cv_folds = min(5, len(X))
        rf_pred  = cross_val_predict(rf,  X, y, cv=cv_folds)
        xgb_pred = cross_val_predict(xgb, X, y, cv=cv_folds)
    rf.fit(X, y)
    xgb.fit(X, y)
    return rf, xgb, rf_pred, xgb_pred


def show_insight(text, icon="💡"):
    if text:
        st.info(f"{icon} {text}")


def make_pdf(df, target_col, time_col, rf, xgb, y, rf_pred, xgb_pred, flags,
             titer_path, feat_path, cmp_path, shap_path, corr_path,
             harvest_path, whatif_path, cellline_path,
             ai_summary, ai_flags, ai_next_run, timestamp):
    pdf = FPDF()
    pdf.set_margins(15, 15, 15)
    pdf.add_page()
    W = pdf.w - 2 * pdf.l_margin

    def heading(txt, size=13):
        pdf.set_font("Helvetica", "B", size)
        pdf.cell(W, 10, txt, ln=True)
        pdf.set_font("Helvetica", "", 11)

    def trow(k, v):
        cw = W / 2
        pdf.cell(cw, 8, str(k), border=1)
        pdf.cell(cw, 8, str(v), border=1, ln=True)

    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(W, 14, "Bioprocess ML Run Report", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(W, 8, f"Generated: {timestamp}", ln=True, align="C")
    pdf.ln(4)

    if ai_summary:
        heading("Executive Summary")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(W, 6, clean_text(ai_summary))
        pdf.ln(3)

    heading("1. Run Metrics")
    rows = {
        "Target Variable":    target_col,
        "Time Column":        time_col,
        "Total Samples":      len(df),
        "Final Target Value": f"{float(df[target_col].iloc[-1]):.4f}",
        "Peak Target Value":  f"{float(df[target_col].max()):.4f}",
        "RF R2 (CV)":         f"{r2_score(y, rf_pred):.4f}",
        "XGB R2 (CV)":        f"{r2_score(y, xgb_pred):.4f}",
        "RF MAE":             f"{mean_absolute_error(y, rf_pred):.4f}",
        "XGB MAE":            f"{mean_absolute_error(y, xgb_pred):.4f}",
    }
    for k, v in rows.items():
        trow(k, v)
    pdf.ln(4)

    heading("2. Process Flags")
    for sev, msg in flags:
        prefix = "[!] " if sev == "warn" else "[OK] "
        pdf.multi_cell(W, 8, clean_text(prefix + msg))
    if ai_flags:
        pdf.ln(2)
        pdf.set_font("Helvetica", "I", 10)
        pdf.multi_cell(W, 6, clean_text(ai_flags))
    pdf.ln(4)

    if ai_next_run:
        heading("3. Recommendations for Next Run")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(W, 6, clean_text(ai_next_run))
        pdf.ln(4)

    heading("4. Target Variable Curve")
    if titer_path and os.path.exists(titer_path):
        pdf.image(titer_path, w=W)
    pdf.ln(4)

    pdf.add_page()
    heading("5. Model Comparison")
    if cmp_path and os.path.exists(cmp_path):
        pdf.image(cmp_path, w=W)
    pdf.ln(4)

    heading("6. Feature Importance")
    if feat_path and os.path.exists(feat_path):
        pdf.image(feat_path, w=W)
    pdf.ln(4)

    if shap_path and os.path.exists(shap_path):
        pdf.add_page()
        heading("7. SHAP Summary")
        pdf.image(shap_path, w=W)

    if corr_path and os.path.exists(corr_path):
        pdf.add_page()
        heading("8. Correlation Heatmap")
        pdf.image(corr_path, w=W)

    if harvest_path and os.path.exists(harvest_path):
        pdf.add_page()
        heading("9. Harvest Timing")
        pdf.image(harvest_path, w=W)

    if whatif_path and os.path.exists(whatif_path):
        pdf.add_page()
        heading("10. Feed Strategy Sensitivity")
        pdf.image(whatif_path, w=W)

    if cellline_path and os.path.exists(cellline_path):
        pdf.add_page()
        heading("11. Cell Line Comparison")
        pdf.image(cellline_path, w=W)

    return bytes(pdf.output())


def make_excel(df, target_col, time_col, rf, xgb, feature_cols, y, rf_pred, xgb_pred, flags, timestamp):
    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1a3a4a")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, ncols):
        for cell in ws[1]:
            if cell.column <= ncols:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "Predictions"
    export_cols = [time_col, target_col] + [c for c in feature_cols if c not in [time_col, target_col]]
    export_cols = [c for c in export_cols if c in df.columns]
    df_out = df[export_cols].copy()
    df_out['RF Predicted']  = rf_pred
    df_out['XGB Predicted'] = xgb_pred
    for ci, cn in enumerate(df_out.columns, 1):
        ws1.cell(1, ci, cn)
    for ri, row in enumerate(df_out.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            ws1.cell(ri, ci, float(v) if isinstance(v, (np.floating, np.integer)) else v)
    style_header(ws1, len(df_out.columns))

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Random Forest", "XGBoost"])
    ws2.append(["R2 (CV)", round(r2_score(y, rf_pred), 4), round(r2_score(y, xgb_pred), 4)])
    ws2.append(["MAE", round(mean_absolute_error(y, rf_pred), 4), round(mean_absolute_error(y, xgb_pred), 4)])
    ws2.append([])
    ws2.append([f"Report generated: {timestamp}"])
    ws2.append(["Process Flags"])
    for sev, msg in flags:
        ws2.append([clean_text(("[!] " if sev == "warn" else "[OK] ") + msg)])
    style_header(ws2, 3)

    ws3 = wb.create_sheet("Feature Importance")
    ws3.append(["Feature", "RF Importance", "Rank"])
    imp = sorted(zip(feature_cols, rf.feature_importances_), key=lambda x: x[1], reverse=True)
    for rank, (feat, val) in enumerate(imp, 1):
        ws3.append([feat, round(float(val), 6), rank])
    style_header(ws3, 3)

    ws4 = wb.create_sheet("Target Chart")
    ws4.append([time_col, "Actual", "RF Predicted", "XGB Predicted"])
    for _, row in df_out.iterrows():
        ws4.append([row[time_col], row[target_col], row['RF Predicted'], row['XGB Predicted']])
    chart = LineChart()
    chart.title = f"{target_col}: Actual vs Predicted"
    chart.y_axis.title = target_col
    data = Reference(ws4, min_col=2, max_col=4, min_row=1, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    ws4.add_chart(chart, "F2")
    style_header(ws4, 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# TABS
# =============================================================================
tab_main, tab_harvest, tab_feedopt, tab_earlywarning, tab_cellline = st.tabs([
    "📊 Main Analysis",
    "🌾 Harvest Timing",
    "🧪 Feed Optimizer",
    "⚡ Early Warning",
    "🔬 Cell Line Comparison"
])

# =============================================================================
# TAB 1 — MAIN ANALYSIS
# =============================================================================
with tab_main:

    st.markdown('<div class="step-card"><strong>📁 Step 1 — Upload Your Data</strong></div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Upload CSV or Excel file", type=["csv","xlsx"], key="main_upload")

    if uploaded:
        try:
            file_bytes = uploaded.read()
            if uploaded.name.endswith('.csv'):
                df = load_csv(file_bytes)
            else:
                xl = pd.ExcelFile(io.BytesIO(file_bytes))
                sheet = st.selectbox("Select sheet", xl.sheet_names)
                df = load_excel(file_bytes, sheet)
            st.session_state.df = df
            st.success(f"Loaded {len(df)} rows x {len(df.columns)} columns")
            with st.expander("📋 Preview Data"):
                st.dataframe(df.head(15), use_container_width=True)
        except Exception as e:
            st.error(f"Error loading file: {e}")

    if st.session_state.df is not None:
        df = st.session_state.df
        all_cols     = df.columns.tolist()
        numeric_cols = df.select_dtypes(include=np.number).columns.tolist()

        st.markdown('<div class="step-card"><strong>⚙️ Step 2 — Map Your Columns</strong></div>', unsafe_allow_html=True)
        auto_target, auto_time, auto_group = smart_detect(all_cols)

        col1, col2, col3 = st.columns(3)
        with col1:
            target_col = st.selectbox("🎯 Target variable", numeric_cols,
                                       index=numeric_cols.index(auto_target) if auto_target in numeric_cols else 0)
        with col2:
            time_col = st.selectbox("📅 Time / Day column", all_cols,
                                     index=all_cols.index(auto_time) if auto_time in all_cols else 0)
        with col3:
            group_options = ["None"] + all_cols
            group_sel = st.selectbox("🔬 Experiment/Run ID (optional)", group_options,
                                      index=group_options.index(auto_group) if auto_group in group_options else 0)
            group_col = None if group_sel == "None" else group_sel

        st.subheader("Select Feature Columns")
        exclude = {target_col, time_col, group_col}
        default_features = [c for c in numeric_cols if c not in exclude]
        feature_cols = st.multiselect("Features", default_features, default=default_features)

        st.session_state.feature_cols = feature_cols
        st.session_state.target_col   = target_col
        st.session_state.time_col     = time_col
        st.session_state.group_col    = group_col

    if st.session_state.feature_cols:
        st.markdown('<div class="step-card"><strong>🤖 Step 3 — Train Models & Generate Insights</strong></div>', unsafe_allow_html=True)

        if st.button("🚀 Run Analysis", type="primary", use_container_width=True):
            df           = st.session_state.df
            target_col   = st.session_state.target_col
            time_col     = st.session_state.time_col
            group_col    = st.session_state.group_col
            feature_cols = st.session_state.feature_cols

            df_model = df[feature_cols + [target_col]].apply(pd.to_numeric, errors='coerce').dropna()
            X = df_model[feature_cols]
            y = df_model[target_col]
            df_plot = df.loc[df_model.index]

            if len(X) < 6:
                st.error(f"Only {len(X)} usable rows - need at least 6.")
                st.stop()

            groups = df_plot[group_col].astype(str) if group_col else None
            st.info(f"Training on {len(X)} rows with {len(feature_cols)} features")

            with st.spinner("Training models..."):
                rf, xgb, rf_pred, xgb_pred = train_models(
                    X.values.tolist(), list(X.columns),
                    y.tolist(), y.name,
                    groups.tolist() if groups is not None else None,
                    n_estimators
                )

            if use_shap:
                with st.spinner("Computing SHAP values..."):
                    shap_values = shap.TreeExplainer(rf).shap_values(X)
                st.session_state.shap_values = shap_values

            flags = generate_flags(df_plot, target_col, time_col, nh3_thresh, glc_thresh, vcd_thresh, lg_thresh)

            st.session_state.rf          = rf
            st.session_state.xgb         = xgb
            st.session_state.y           = y
            st.session_state.rf_pred     = rf_pred
            st.session_state.xgb_pred    = xgb_pred
            st.session_state.X_model     = X
            st.session_state.df_plot     = df_plot
            st.session_state.flags       = flags

            with st.spinner("Generating AI interpretations..."):
                run_all_ai_interpretations(df_plot, target_col, time_col, feature_cols,
                                           rf, y, rf_pred, xgb_pred, flags, group_col)

            st.success("Analysis and AI interpretations complete!")

    if st.session_state.rf is not None:
        rf           = st.session_state.rf
        xgb          = st.session_state.xgb
        y            = st.session_state.y
        rf_pred      = st.session_state.rf_pred
        xgb_pred     = st.session_state.xgb_pred
        flags        = st.session_state.flags
        df_plot      = st.session_state.df_plot
        X_model      = st.session_state.X_model
        shap_values  = st.session_state.shap_values
        target_col   = st.session_state.target_col
        time_col     = st.session_state.time_col
        group_col    = st.session_state.group_col
        feature_cols = st.session_state.feature_cols

        st.markdown('<div class="step-card"><strong>📊 Step 4 — Results</strong></div>', unsafe_allow_html=True)

        # AI Run Summary — shown first, prominent
        if st.session_state.get('ai_error'):
            st.warning(f"AI interpretation unavailable: {st.session_state['ai_error']}")
        if st.session_state.ai_summary:
            st.subheader("📝 AI Run Summary")
            st.info(st.session_state.ai_summary)

        # Metrics
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("RF R2 (CV)",  f"{r2_score(y, rf_pred):.4f}")
        c2.metric("XGB R2 (CV)", f"{r2_score(y, xgb_pred):.4f}")
        c3.metric("RF MAE",      f"{mean_absolute_error(y, rf_pred):.4f}")
        c4.metric("XGB MAE",     f"{mean_absolute_error(y, xgb_pred):.4f}")

        # Process flags with AI explanations
        st.subheader("🚩 Process Flags")
        for sev, msg in flags:
            if sev == "warn":
                st.warning(f"⚠️ {msg}")
            else:
                st.success(f"✅ {msg}")
        if st.session_state.ai_flags:
            with st.expander("💬 What do these flags mean in plain English?", expanded=True):
                st.write(st.session_state.ai_flags)

        titer_path = feat_path = cmp_path = shap_path = corr_path = None

        # Titer curve
        st.subheader(f"📈 {target_col} Over Time")
        if group_col and group_col in df_plot.columns:
            all_batches = sorted(df_plot[group_col].astype(str).unique().tolist())
            selected_batches = st.multiselect("Filter by batch/run", all_batches, default=all_batches, key="batch_filter")
            mask = df_plot[group_col].astype(str).isin(selected_batches) if selected_batches else pd.Series([True]*len(df_plot), index=df_plot.index)
        else:
            mask = pd.Series([True]*len(df_plot), index=df_plot.index)

        df_f  = df_plot[mask]
        y_f   = y[mask]
        rf_f  = rf_pred[mask.values]
        xgb_f = xgb_pred[mask.values]

        fig1, ax1 = plt.subplots(figsize=(12, 4))
        tab_colors = plt.cm.tab10.colors
        if group_col and group_col in df_f.columns:
            from matplotlib.lines import Line2D
            for i, batch in enumerate(df_f[group_col].astype(str).unique()):
                b_idx = df_f[df_f[group_col].astype(str)==batch].index
                bm = df_plot.index.isin(b_idx)
                ax1.plot(df_f.loc[b_idx, time_col], y[bm],       'o-',  color=tab_colors[i%10], label=str(batch))
                ax1.plot(df_f.loc[b_idx, time_col], rf_pred[bm], 's--', color=tab_colors[i%10], alpha=0.55)
                ax1.plot(df_f.loc[b_idx, time_col], xgb_pred[bm],'^:',  color=tab_colors[i%10], alpha=0.35)
            style_legend = [Line2D([0],[0],color='gray',marker='o',ls='-',label='Actual'),
                            Line2D([0],[0],color='gray',marker='s',ls='--',label='RF Pred',alpha=0.6),
                            Line2D([0],[0],color='gray',marker='^',ls=':',label='XGB Pred',alpha=0.4)]
            l1 = ax1.legend(handles=style_legend, loc='upper left', fontsize=8, title='Line Style')
            ax1.add_artist(l1)
            ax1.legend(loc='lower right', fontsize=8, title='Batch/Run')
        else:
            ax1.plot(df_f[time_col].values, y_f.values,'o-', label='Actual', color='steelblue')
            ax1.plot(df_f[time_col].values, rf_f,'s--', label='RF Predicted', color='coral')
            ax1.plot(df_f[time_col].values, xgb_f,'^:', label='XGB Predicted', color='seagreen')
            ax1.legend(fontsize=9)
        ax1.set_xlabel(time_col); ax1.set_ylabel(target_col)
        ax1.set_title(f"{target_col} - Actual vs Predicted", fontsize=13)
        ax1.grid(True, alpha=0.3); plt.tight_layout()
        st.pyplot(fig1, use_container_width=True)
        titer_path = tempfile.mktemp(suffix=".png")
        fig1.savefig(titer_path, bbox_inches='tight', dpi=150); plt.close()
        show_insight(st.session_state.ai_titer, "📈")

        col_a, col_b = st.columns(2)
        fig_h = max(5, len(feature_cols) * 0.45)

        with col_a:
            st.subheader("🔍 Feature Importance (RF)")
            fig2, ax2 = plt.subplots(figsize=(7, fig_h))
            imp_df = pd.DataFrame({'Feature': feature_cols, 'Importance': rf.feature_importances_}).sort_values('Importance')
            bars = ax2.barh(imp_df['Feature'], imp_df['Importance'], color='steelblue', height=0.6)
            for bar, val in zip(bars, imp_df['Importance']):
                ax2.text(bar.get_width()+0.001, bar.get_y()+bar.get_height()/2, f'{val:.3f}', va='center', fontsize=8)
            ax2.set_xlabel('Importance Score'); ax2.set_title('RF Feature Importance')
            ax2.tick_params(axis='y', labelsize=9); plt.tight_layout()
            st.pyplot(fig2, use_container_width=True)
            feat_path = tempfile.mktemp(suffix=".png")
            fig2.savefig(feat_path, bbox_inches='tight', dpi=150); plt.close()
            show_insight(st.session_state.ai_feat, "🔍")

        with col_b:
            if shap_values is not None:
                st.subheader("🧠 SHAP Summary Plot")
                shap.summary_plot(shap_values, X_model, plot_type="bar", feature_names=feature_cols, show=False)
                fig3 = plt.gcf(); fig3.set_size_inches(7, fig_h)
                plt.title('SHAP Feature Impact'); plt.tight_layout()
                st.pyplot(fig3, use_container_width=True)
                shap_path = tempfile.mktemp(suffix=".png")
                fig3.savefig(shap_path, bbox_inches='tight', dpi=150); plt.close()
                show_insight(st.session_state.ai_shap, "🧠")
            else:
                st.info("Enable SHAP in the sidebar.")

        st.subheader("🆚 Model Comparison")
        fig4, (ax4a, ax4b) = plt.subplots(1, 2, figsize=(12, 4))
        for ax, pred, name, color in [(ax4a,rf_pred,"Random Forest","coral"),(ax4b,xgb_pred,"XGBoost","seagreen")]:
            ax.scatter(y, pred, alpha=0.7, color=color, edgecolors='white', linewidths=0.5, s=50)
            mn, mx = float(y.min()), float(y.max())
            ax.plot([mn,mx],[mn,mx],'k--',lw=1.5,label='Perfect fit')
            ax.set_xlabel(f'Actual {target_col}'); ax.set_ylabel(f'Predicted {target_col}')
            ax.set_title(f'{name}  R2={r2_score(y,pred):.3f}  MAE={mean_absolute_error(y,pred):.3f}')
            ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
        plt.tight_layout()
        st.pyplot(fig4, use_container_width=True)
        cmp_path = tempfile.mktemp(suffix=".png")
        fig4.savefig(cmp_path, bbox_inches='tight', dpi=150); plt.close()

        st.subheader("🔗 Correlation Heatmap")
        corr_cols = feature_cols + [target_col]
        corr_df   = df_plot[corr_cols].select_dtypes(include=np.number).corr()
        fig5_h = max(6, len(corr_cols)*0.6)
        fig5_w = min(14, len(corr_cols)*0.9+2)
        fig5, ax5 = plt.subplots(figsize=(fig5_w, fig5_h))
        mask_tri = np.triu(np.ones_like(corr_df, dtype=bool))
        sns.heatmap(corr_df, mask=mask_tri, annot=True, fmt=".2f", cmap="RdYlBu_r", center=0,
                    ax=ax5, linewidths=0.5, annot_kws={"size":8}, cbar_kws={"shrink":0.8})
        ax5.set_title("Feature Correlation Matrix", fontsize=13); plt.tight_layout()
        st.pyplot(fig5, use_container_width=True)
        corr_path = tempfile.mktemp(suffix=".png")
        fig5.savefig(corr_path, bbox_inches='tight', dpi=150); plt.close()
        show_insight(st.session_state.ai_corr, "🔗")

        # Next run recommendations
        if st.session_state.ai_next_run:
            st.subheader("🎯 Recommendations for Your Next Run")
            st.info(st.session_state.ai_next_run)

        # Downloads
        st.markdown('<div class="step-card"><strong>📥 Step 5 — Download Reports</strong></div>', unsafe_allow_html=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        ts_label  = datetime.now().strftime("%Y-%m-%d %H:%M")
        d1, d2 = st.columns(2)
        with d1:
            try:
                pdf_bytes = make_pdf(
                    df_plot, target_col, time_col, rf, xgb, y, rf_pred, xgb_pred, flags,
                    titer_path, feat_path, cmp_path, shap_path, corr_path,
                    None, None, None,
                    st.session_state.ai_summary, st.session_state.ai_flags,
                    st.session_state.ai_next_run, ts_label
                )
                st.download_button("📄 Download PDF Report", data=pdf_bytes,
                                   file_name=f"bioprocess_report_{timestamp}.pdf",
                                   mime="application/pdf", use_container_width=True)
            except Exception as e:
                st.error(f"PDF error: {e}")
        with d2:
            try:
                excel_bytes = make_excel(df_plot, target_col, time_col, rf, xgb,
                                         feature_cols, y, rf_pred, xgb_pred, flags, ts_label)
                st.download_button("📊 Download Excel Report", data=excel_bytes,
                                   file_name=f"bioprocess_report_{timestamp}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            except Exception as e:
                st.error(f"Excel error: {e}")
    else:
        st.info("👆 Upload a file and complete Steps 1-3 to see results.")


# =============================================================================
# TAB 2 — HARVEST TIMING
# =============================================================================
with tab_harvest:
    st.markdown("## 🌾 Harvest Timing Predictor")
    st.markdown("Predicts the optimal harvest day based on your historical run data.")

    if st.session_state.rf is None:
        st.info("Run the Main Analysis first.")
    else:
        rf           = st.session_state.rf
        df_plot      = st.session_state.df_plot
        target_col   = st.session_state.target_col
        time_col     = st.session_state.time_col
        feature_cols = st.session_state.feature_cols
        numeric_feats = [c for c in feature_cols if c in df_plot.select_dtypes(include=np.number).columns]

        if numeric_feats:
            last_row   = df_plot[numeric_feats].iloc[-1].copy()
            max_day    = float(df_plot[time_col].max())
            extra_days = st.slider("Simulate days beyond current run end", 1, 30, 10)
            time_vals  = np.arange(max_day + 1, max_day + extra_days + 1)
            sim_rows   = pd.DataFrame([last_row.values] * len(time_vals), columns=numeric_feats)
            valid_sim_cols = [c for c in feature_cols if c in sim_rows.columns]
            if time_col in feature_cols and time_col in sim_rows.columns:
                sim_rows[time_col] = time_vals
            sim_preds  = rf.predict(sim_rows[valid_sim_cols])

            hist_times = df_plot[time_col].values
            hist_preds = st.session_state.rf_pred
            all_times  = np.concatenate([hist_times, time_vals])
            all_preds  = np.concatenate([hist_preds, sim_preds])
            best_idx   = np.argmax(all_preds)
            best_day   = all_times[best_idx]
            best_pred  = all_preds[best_idx]

            col1, col2, col3 = st.columns(3)
            col1.metric("Predicted Optimal Harvest Day", f"{best_day:.0f}")
            col2.metric("Predicted Peak Titer",          f"{best_pred:.2f}")
            col3.metric("Current Run End Day",            f"{max_day:.0f}")

            fig_h, ax_h = plt.subplots(figsize=(12, 4))
            ax_h.plot(hist_times, st.session_state.y.values,'o-', color='steelblue', label='Actual (historical)', zorder=3)
            ax_h.plot(hist_times, hist_preds,'s--', color='coral', label='RF Predicted (historical)', alpha=0.7)
            ax_h.plot(time_vals, sim_preds, '^:', color='seagreen', label='Simulated future', alpha=0.8)
            ax_h.axvline(best_day, color='gold', lw=2, ls='--', label=f'Optimal harvest: Day {best_day:.0f}')
            ax_h.axvline(max_day,  color='gray', lw=1, ls=':', alpha=0.6, label='Current run end')
            ax_h.set_xlabel(time_col); ax_h.set_ylabel(target_col)
            ax_h.set_title("Harvest Timing Prediction"); ax_h.legend(fontsize=8); ax_h.grid(True, alpha=0.3)
            plt.tight_layout()
            st.pyplot(fig_h, use_container_width=True)
            plt.close()

            # Auto AI interpretation
            harvest_insight = ai_interpret(f"""You are a bioprocess scientist. The model predicts optimal harvest at day {best_day:.0f} with predicted titer {best_pred:.2f}. The run currently ends at day {max_day:.0f}.
In 2 sentences, give a plain-English harvest recommendation. Should they extend the run? What is the risk of waiting?""")
            show_insight(harvest_insight, "🌾")


# =============================================================================
# TAB 3 — FEED OPTIMIZER
# =============================================================================
with tab_feedopt:
    st.markdown("## 🧪 Feed Strategy Optimizer")
    st.markdown("Adjust metabolite levels to simulate what the titer would be under different feed conditions.")

    if st.session_state.rf is None:
        st.info("Run the Main Analysis first.")
    else:
        rf           = st.session_state.rf
        df_plot      = st.session_state.df_plot
        target_col   = st.session_state.target_col
        feature_cols = st.session_state.feature_cols
        time_col     = st.session_state.time_col
        numeric_feats = [c for c in feature_cols if c in df_plot.select_dtypes(include=np.number).columns]

        if numeric_feats:
            slider_vals = {}
            cols = st.columns(min(3, len(numeric_feats)))
            for i, feat in enumerate(numeric_feats):
                mn  = float(df_plot[feat].min())
                mx  = float(df_plot[feat].max())
                avg = float(df_plot[feat].mean())
                rng = mx - mn
                step = rng / 100 if rng > 0 else 0.01
                with cols[i % len(cols)]:
                    slider_vals[feat] = st.slider(feat, mn, mx, avg, step, key=f"wi_{feat}")

            input_row     = pd.DataFrame([[slider_vals[f] for f in numeric_feats]], columns=numeric_feats)
            valid_cols    = [c for c in feature_cols if c in input_row.columns]
            whatif_pred   = rf.predict(input_row[valid_cols])[0]
            baseline_row  = pd.DataFrame([[float(df_plot[f].mean()) for f in numeric_feats]], columns=numeric_feats)
            baseline_pred = rf.predict(baseline_row[valid_cols])[0]
            delta         = whatif_pred - baseline_pred

            st.markdown("---")
            c1, c2, c3 = st.columns(3)
            c1.metric("Predicted Titer (your settings)", f"{whatif_pred:.2f}")
            c2.metric("Baseline Titer (mean conditions)", f"{baseline_pred:.2f}")
            c3.metric("Delta vs Baseline", f"{delta:+.2f}", delta_color="normal")

            # Sensitivity
            st.subheader("Feature Sensitivity")
            sensitivity = {}
            for feat in numeric_feats:
                sweep = np.linspace(float(df_plot[feat].min()), float(df_plot[feat].max()), 20)
                preds = []
                for val in sweep:
                    row = input_row.copy(); row[feat] = val
                    preds.append(rf.predict(row[valid_cols])[0])
                sensitivity[feat] = max(preds) - min(preds)

            sens_df = pd.DataFrame(list(sensitivity.items()), columns=['Feature','Sensitivity']).sort_values('Sensitivity')
            fig_wi, ax_wi = plt.subplots(figsize=(8, max(4, len(numeric_feats)*0.4)))
            ax_wi.barh(sens_df['Feature'], sens_df['Sensitivity'], color='teal', height=0.6)
            ax_wi.set_xlabel(f'Predicted {target_col} range'); ax_wi.set_title('Feature Sensitivity')
            plt.tight_layout()
            st.pyplot(fig_wi, use_container_width=True)
            plt.close()

            top_sensitive = sens_df.sort_values('Sensitivity', ascending=False).head(3)
            feed_insight = ai_interpret(f"""You are a bioprocess scientist. 
The what-if titer prediction is {whatif_pred:.2f} vs baseline {baseline_pred:.2f} (delta: {delta:+.2f}).
Top 3 most sensitive features: {', '.join(top_sensitive['Feature'].tolist())}.
Current slider values: {', '.join([f'{k}: {v:.2f}' for k, v in slider_vals.items()])}.
In 2-3 sentences, give a plain-English feed strategy recommendation. Focus on which metabolites to adjust and why.""")
            show_insight(feed_insight, "🧪")


# =============================================================================
# TAB 4 — EARLY WARNING
# =============================================================================
with tab_earlywarning:
    st.markdown("## ⚡ Early Warning System")
    st.markdown("Upload early timepoint data from a new run to forecast its final titer against your historical average.")

    if st.session_state.rf is None:
        st.info("Run the Main Analysis first.")
    else:
        rf           = st.session_state.rf
        target_col   = st.session_state.target_col
        time_col     = st.session_state.time_col
        feature_cols = st.session_state.feature_cols
        df_plot      = st.session_state.df_plot
        group_col    = st.session_state.group_col

        new_upload = st.file_uploader("Upload early timepoint data", type=["csv","xlsx"], key="ew_upload")

        if new_upload:
            try:
                nb = new_upload.read()
                if new_upload.name.endswith('.csv'):
                    df_new = load_csv(nb)
                else:
                    xl2 = pd.ExcelFile(io.BytesIO(nb))
                    sh2 = st.selectbox("Select sheet", xl2.sheet_names, key="ew_sheet")
                    df_new = load_excel(nb, sh2)

                st.success(f"Loaded {len(df_new)} early timepoint rows")
                st.dataframe(df_new.head(), use_container_width=True)

                X_new = pd.DataFrame()
                for feat in feature_cols:
                    if feat in df_new.columns:
                        X_new[feat] = pd.to_numeric(df_new[feat], errors='coerce')
                    else:
                        X_new[feat] = float(df_plot[feat].mean()) if feat in df_plot.columns else 0.0
                X_new = X_new.fillna(X_new.mean())
                preds_new = rf.predict(X_new)

                final_forecast = preds_new[-1]
                peak_forecast  = preds_new.max()
                day_col_vals   = df_new[time_col].values if time_col in df_new.columns else np.arange(len(preds_new))

                hist_finals = []
                if group_col and group_col in df_plot.columns:
                    for grp in df_plot[group_col].unique():
                        gdf = df_plot[df_plot[group_col]==grp]
                        hist_finals.append(float(gdf[target_col].iloc[-1]))
                else:
                    hist_finals = [float(df_plot[target_col].iloc[-1])]

                hist_mean = np.mean(hist_finals)
                hist_std  = np.std(hist_finals) if len(hist_finals) > 1 else 0
                z = (final_forecast - hist_mean) / hist_std if hist_std > 0 else 0

                col1, col2, col3 = st.columns(3)
                col1.metric("Forecasted Final Titer", f"{final_forecast:.2f}")
                col2.metric("Historical Mean",         f"{hist_mean:.2f}")
                col3.metric("Z-score vs Historical",   f"{z:.2f}")

                fig_ew, ax_ew = plt.subplots(figsize=(10, 4))
                ax_ew.plot(day_col_vals, preds_new, 'o-', color='steelblue', label='Forecast')
                ax_ew.axhline(hist_mean, color='gray', ls='--', lw=1.5, label=f'Historical mean ({hist_mean:.2f})')
                if hist_std > 0:
                    ax_ew.axhspan(hist_mean-hist_std, hist_mean+hist_std, alpha=0.15, color='gray', label='Historical +/-1 std')
                ax_ew.set_xlabel(time_col); ax_ew.set_ylabel(f'Predicted {target_col}')
                ax_ew.set_title('Early Warning: Forecasted Titer Trajectory')
                ax_ew.legend(fontsize=9); ax_ew.grid(True, alpha=0.3)
                plt.tight_layout()
                st.pyplot(fig_ew, use_container_width=True)
                plt.close()

                ew_insight = ai_interpret(f"""You are a bioprocess scientist reviewing an early run forecast.
Early data covers {len(df_new)} samples. Forecasted final titer: {final_forecast:.2f}.
Historical mean: {hist_mean:.2f}, std: {hist_std:.2f}, z-score: {z:.2f}.
In 2-3 sentences, give a plain-English early warning assessment. Should the scientist intervene? What specifically should they check?""")
                show_insight(ew_insight, "⚡")

            except Exception as e:
                st.error(f"Error: {e}")
        else:
            st.markdown("""
**How to use:** Run the Main Analysis on completed historical runs first, then upload just the first few days of a new run here. The model will forecast whether the run is on track.
            """)


# =============================================================================
# TAB 5 — CELL LINE COMPARISON
# =============================================================================
with tab_cellline:
    st.markdown("## 🔬 Cell Line Comparison")
    st.markdown("Benchmarks all runs or cell lines in your dataset against each other.")

    if st.session_state.rf is None:
        st.info("Run the Main Analysis first.")
    else:
        df_plot      = st.session_state.df_plot
        target_col   = st.session_state.target_col
        time_col     = st.session_state.time_col
        group_col    = st.session_state.group_col
        rf_pred      = st.session_state.rf_pred
        feature_cols = st.session_state.feature_cols

        if group_col and group_col in df_plot.columns:
            groups = df_plot[group_col].astype(str).unique()
            summary_rows = []
            for grp in groups:
                gdf   = df_plot[df_plot[group_col].astype(str)==grp]
                gmask = df_plot[group_col].astype(str)==grp
                gpred = rf_pred[gmask.values]
                gact  = st.session_state.y[gmask]
                row = {
                    "Run / Cell Line": grp,
                    "Samples":         len(gdf),
                    "Duration":        int(gdf[time_col].max()) if time_col in gdf.columns else "N/A",
                    "Final Titer":     round(float(gdf[target_col].iloc[-1]), 2),
                    "Peak Titer":      round(float(gdf[target_col].max()), 2),
                    "Mean Titer":      round(float(gdf[target_col].mean()), 2),
                }
                if len(gact) > 1:
                    row["R2"] = round(r2_score(gact, gpred), 3)
                summary_rows.append(row)

            summary_df = pd.DataFrame(summary_rows).sort_values("Peak Titer", ascending=False)
            st.dataframe(summary_df, use_container_width=True)

            fig_cl, axes = plt.subplots(1, 3, figsize=(14, 5))
            colors = plt.cm.tab10.colors
            for ax, metric in zip(axes, ["Final Titer","Peak Titer","Mean Titer"]):
                vals = summary_df[metric].values
                labs = summary_df["Run / Cell Line"].values
                bars = ax.bar(labs, vals, color=[colors[i%10] for i in range(len(labs))])
                for bar, val in zip(bars, vals):
                    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+max(vals)*0.01,
                            f'{val:.1f}', ha='center', fontsize=8)
                ax.set_title(metric); ax.set_ylabel(target_col)
                ax.tick_params(axis='x', rotation=45); ax.grid(True, alpha=0.2, axis='y')
            plt.suptitle("Cell Line / Run Comparison", fontsize=14, y=1.02)
            plt.tight_layout()
            st.pyplot(fig_cl, use_container_width=True)
            plt.close()

            # Metabolic fingerprint
            st.subheader("Metabolic Fingerprint")
            numeric_feats = [c for c in feature_cols if c in df_plot.select_dtypes(include=np.number).columns]
            if len(numeric_feats) >= 3:
                means_df = df_plot.groupby(group_col)[numeric_feats].mean()
                norm_df  = (means_df - means_df.min()) / (means_df.max() - means_df.min() + 1e-9)
                fig_fp, ax_fp = plt.subplots(figsize=(12, 5))
                sns.heatmap(norm_df.T, annot=True, fmt=".2f", cmap="YlOrRd",
                            ax=ax_fp, linewidths=0.5, annot_kws={"size":8},
                            cbar_kws={"label":"Normalised mean (0-1)"})
                ax_fp.set_title("Metabolic Fingerprint by Run (normalised)")
                plt.tight_layout()
                st.pyplot(fig_fp, use_container_width=True)
                plt.close()

            best_run    = summary_df.iloc[0]["Run / Cell Line"]
            worst_run   = summary_df.iloc[-1]["Run / Cell Line"]
            best_titer  = summary_df.iloc[0]["Peak Titer"]
            worst_titer = summary_df.iloc[-1]["Peak Titer"]
            cl_insight = ai_interpret(f"""You are a bioprocess scientist comparing {len(groups)} runs.
Best run: {best_run} with peak titer {best_titer:.2f}. Worst run: {worst_run} with peak titer {worst_titer:.2f}.
All runs: {summary_df[['Run / Cell Line','Final Titer','Peak Titer']].to_string(index=False)}.
In 3 sentences, give a plain-English comparison. Which cell line or run should be prioritised and why?""")
            show_insight(cl_insight, "🔬")

        else:
            st.markdown('<div class="warning-box">No group/run ID column was selected in the Main Analysis. Go back to Step 2 and set the Experiment/Run ID column to enable multi-run comparison.</div>', unsafe_allow_html=True)
